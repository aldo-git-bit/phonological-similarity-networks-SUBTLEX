import networkx as nx
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import math
import eng_to_ipa as ipa
import textdistance
import spacy
import os



class Words:
    def __init__(self, WORD, IPA, IPA_LIST):
        self.WORD = WORD
        self.IPA = IPA
        self.IPA_LIST = IPA_LIST


def validateWordEdgeList(edgelist):
    irregular_words = {
        "person":["people"],
        "people":["person"],
        "ox":["oxen"],
        "oxen":["ox"],
        "man":["men"],
        "men":["man"],
        "woman":["women"],
        "women":["woman"],
        "caveman":["cavemen"],
        "cavemen":["caveman"],
        "policeman":["policemen"],
        "policemen":["policeman"],
        "child":["children"],
        "children":["child"],
        "foot":["feet"],
        "feet":["foot"],
        "goose":["geese"],
        "geese":["goose"],
        "mouse":["mice"],
        "mice":["mouse"],
        "louse":["lice"],
        "lice":["louse"],
        "penny":["pence"],
        "pence":["penny"],
        "abide":["abided, abode"],
        "abided":["abide", "abode"],
        "abode":["abide","abided"],
        "arise":["arose","arisen"],
        "arose":["arise","arisen"],
        "arisen":["arise","arose"],
        "awake":["awaked","awoke","awoken"],
        "awaked":["awake","awoke","awoken"],
        "awoke":["awake","awaked","awoken"],
        "awoken":["awake","awaked","awoke"],
        "be":["was","were","been"],
        "was":["be","were","been"],
        "were":["be","was","been"],
        "been":["be","were","was"],
        "bear":["bore","born"],
        "bore":["bear","born"],
        "born":["bear","bore"],
        "beat":["beaten"],
        "beaten":["beat"],
        "become":["became"],
        "became":["become"],
        "befall": ["befell", "befallen"],
        "befell":["befall","befallen"],
        "befallen":["befall","befell"],
        "beget":["begot","begotten"],
        "begot":["beget","begotten"],
        "begotten":["begot","beget"],
        "begin":["began","begun"],
        "began":["begin","begun"],
        "begun":["begin","began"],
        "behold":["beheld"],
        "beheld":["behold"],
        "bend":["bent"],
        "bent":["bend"],
        "bereave":["bereft"],
        "bereft":["bereave"],
        "beseech":["besought"],
        "besought":["beseech"],
        "beset":["beset"],
        "bespeak":["bespoke","bespoken"],
        "bespoke":["bespeak","bespoken"],
        "bespoken":["bespeak","bespoke"],
        "bestride":["bestrode","bestridden"],
        "bestrode":["bestride","bestridden"],
        "bestridden":["bestride","bestrode"],
        "bet":["bet"],
        "bid":["bade","bidden"],
        "bade":["bid","bidden"],
        "bidden":["bade","bid"],
        "bind":["bound"],
        "bound":["bind"],
        "bite":["bit", "bitten"],
        "bit":["bite","bitten"],
        "bitten":["bite","bit"],
        "bleed":["bled"],
        "bled":["bleed"],
        "blow":["blew","blown"],
        "blew":["blow","blown"],
        "blown":["blew","blow"],
        "break":["broke","broken"],
        "broke":["break","broken"],
        "broken":["break","broke"],
        "breed":["bred"],
        "bred":["breed"],
        "bring":["brought"],
        "brought":["bring"],
        "broadcasted":["broadcast"],
        "broadcast":["broadcasted"],
        "build":["built"],
        "built":["build"],
        "burst":["burst"],
        "burn":["burnt","burned"],
        "burnt":["burn","burned"],
        "burned":["burn","burnt"],
        "buy":["bought"],
        "bought":["buy"],
        "can":["could"],
        "could":["can"],
        "cast": ["cast"],
        "catch":["caught"],
        "caught":["catch"],
        "choose":["chose", "chosen"],
        "chose":["choose","chosen"],
        "chosen":["choose","chose"],
        "cling":["clung"],
        "clung":["cling"],
        "come":["came"],
        "came":["come"],
        "cost":["cost"],
        "creep":["crept"],
        "crept":["creep"],
        "cost":["cost"],
        "deal":["dealt"],
        "dealt":["deal"],
        "dig":["dug"],
        "dug":["dig"],
        "do":["did, done"],
        "did":["do","done"],
        "done":["do","did"],
        "draw":["drew","drawn"],
        "drawn":["draw","drew"],
        "drew":["draw","drawn"],
        "dream":["dreamed","dreamt"],
        "dreamed":["dream","dreamt"],
        "dreamt":["dreamed","dream"],
        "drink":["drank","drunk"],
        "drank":["drink","drunk"],
        "drunk":["drink","drank"],
        "drive":["drove","driven"],
        "drove":["drive","driven"],
        "driven":["drive","drove"],
        "dwell":["dwelled","dwelt"],
        "dwelled":["dwell","dwelt"],
        "dwelt":["dwell","dwelled"],
        "eat":["eaten","ate"],
        "eaten":["eat","ate"],
        "ate":["eaten","eat"],
        "interweave":["interwoven"],
        "interwoven":["interweave"],
        "fall":["fell","fallen"],
        "fell":["fall","fallen"],
        "fallen":["fell","fall"],
        "feed":["fed"],
        "fed":["feed"],
        "feel":["felt"],
        "felt":["feel"],
        "fight":["fought"],
        "fought":["fight"],
        "find":["found"],
        "found":["find"],
        "flee":["fled"],
        "fled":["flee"],
        "fling":["flung"],
        "flung":["fling"],
        "fly":["flew","flown"],
        "flew":["fly","flown"],
        "flown":["fly","flew"],
        "forbid":["forbad","forbade","forbidden"],
        "forbad":["forbid","forbade","forbidden"],
        "forbade":["forbid","forbad","forbidden"],
        "forbidden":["forbid","forbad","forbade"],
        "forecast":["forecasted"],
        "forecasted":["forecast"],
        "forget":["forgot","forgotten"],
        "forgot":["forget","forgotten"],
        "forgotten":["forget","forgot"],
        "forgive":["forgave","forgiven"],
        "forgave":["forgive","forgiven"],
        "forgiven":["forgive","forgave"],
        "forsake":["forsaken","forsook"],
        "forsaken":["forsake","forsook"],
        "forsook":["forsake","forsaken"],
        "foresee":["foresaw","foreseen"],
        "foresaw":["foresee","foreseen"],
        "foreseen":["foresaw","foresee"],
        "foretell":["foretold"],
        "foretold":["foretell"],
        "freeze":["froze","frozen"],
        "frozen":["froze","freeze"],
        "froze":["freeze","frozen"],
        "get":["got"],
        "got":["get"],
        "give":["gave","given"],
        "gave":["give","given"],
        "given":["give","gave"],
        "go":["gone","went"],
        "gone":["go","went"],
        "went":["go","gone"],
        "grind":["ground"],
        "ground":["grind"],
        "grow":["grew","grown"],
        "grew":["grow","grown"],
        "grown":["grew","grow"],
        "hang":["hanged","hung"],
        "hung":["hanged","hang"],
        "hanged":["hung","hang"],
        "have":["had"],
        "had":["have"],
        "hear":["heard"],
        "heard":["hear"],
        "hide":["hid","hidden"],
        "hid":["hide","hidden"],
        "hidden":["hid","hide"],
        "hid":["hid"],
        "hold":["held"],
        "held":["hold"],
        "hurt":["hurt"],
        "keep":["kept"],
        "kept":["keep"],
        "kneel":["knelt"],
        "knelt":["kneel"],
        "know":["knew","known"],
        "known":["knew","know"],
        "knew":["know","known"],
        "know":["known","knew"],
        "lay":["laid"],
        "laid":["lay"],
        "lead":["led"],
        "led":["lead"],
        "lean":["leaned","leant"],
        "leaned":["lean","leant"],
        "leant":["lean","leaned"],
        "leap":["leapt"],
        "leapt":["leap"],
        "learn":["learned","learnt"],
        "learnt":["learned","learn"],
        "learned":["learnt","learn"],
        "left":["leave"],
        "leave":["left"],
        "lent":["lend"],
        "lend":["lent"],
        "let":["let"],
        "lie":["lay","lain"],
        "lay":["lie","lain"],
        "lain":["lie","lay"],
        "lose":["lost"],
        "lost":["lose"],
        "make":["made"],
        "made":["make"],
        "mean":["meant"],
        "meant":["mean"],
        "pay":["paid"],
        "paid":["pay"],
        "mistake":["mistook","mistaken"],
        "mistook":["mistake","mistaken"],
        "mistaken":["mistake","mistook"],
        "overhear":["overheard"],
        "overheard":["overhear"],
        "oversleep":["overslept"],
        "overslept":["oversleep"],
        "rend":["rended","rent"],
        "rended":["rend","rent"],
        "rent":["rended","rend"],
        "ride":["rode","ridden"],
        "rode":["ride","ridden"],
        "ridden":["ride","rode"],
        "ring":["rang","rung"],
        "rang":["ring","rung"],
        "rung":["ring","rang"],
        "rise":["rose","risen"],
        "rose":["rise","risen"],
        "risen":["rise","rose"],
        "run":["ran"],
        "ran":["run"],
        "say":["said"],
        "said":["say"],
        "see":["saw","seen"],
        "saw":["see","seen"],
        "seen":["see","saw"],
        "seek":["sought"],
        "sought":["seek"],
        "sell":["sold"],
        "sold":["sell"],
        "shake":["shook","shaken"],
        "shook":["shake","shaken"],
        "shaken":["shook","shake"],
        "shine":["stone"],
        "stone":["shine"],
        "shit":["shat"],
        "shat":["shit"],
        "shoot":["shot"],
        "shot":["shoot"],
        "show":["showed","shown"],
        "showed":["show","shown"],
        "shown":["show","showed"],
        "shrink":["shrank", "shrunk"],
        "shrank":["shrink","shrunk"],
        "shrunk":["shrank","shrunk"],
        "shrive":["shrove","shriven"],
        "shrove":["shrive","shriven"],
        "shrive":["shrove","shriven"],
        "shriven":["shrove","shrive"],
        "sing":["sang","sung"],
        "sang":["sing","sung"],
        "sung":["sing","sang"],
        "sink":["sank", "sunk"],
        "sank":["sink", "sunk"],
        "sunk":["sink","sank"],
        "sit":["sat"],
        "sat":["sit"],
        "slay":["slew","slain"],
        "slew":["slay","slain"],
        "slain":["slay","slew"],
        "sleep":["slept"],
        "slept":["sleep"],
        "slide":["slid"],
        "slid":["slide"],
        "sling":["slung"],
        "slung":["sling"],
        "slink":["slunk"],
        "slunk":["slink"],
        "smell":["smelled","smelt"],
        "smelled":["smell","smelt"],
        "smelt":["smell","smelled"],
        "smite":["smote","smitten"],
        "smote":["smite","smitten"],
        "smitten":["smite","smote"],
        "speak":["spoke","spoken"],
        "spoke":["speak","spoken"],
        "spoken":["speak","spoke"],
        "speed":["speeded","sped"],
        "speeded":["speed","sped"],
        "sped":["speed","speeded"],
        "spin":["spun"],
        "spun":["spin"],
        "spit":["spat"],
        "spat":["spit"],
        "spin":["spun"],
        "spun":["spin"],
        "spoil":["spoiled","spoilt"],
        "spoiled":["spoil","spoilt"],
        "spoilt":["spoiled","spoil"],
        "spring":["sprang","sprung"],
        "sprang":["spring","sprung"],
        "sprung":["spring","sprang"],
        "stand":["stood"],
        "stood":["stand"],
        "steal":["stole","stolen"],
        "stole":["steal","stolen"],
        "stolen":["stole","steal"],
        "stick":["stuck"],
        "stuck":["stick"],
        "sting":["stung"],
        "stung":["sting"],
        "stink":["stank","stunk"],
        "stank":["stink","stunk"],
        "stunk":["stank","stink"],
        "stride":["strode","stridden"],
        "strode":["stride","stridden"],
        "stridden":["stride","strode"],
        "strike":["struck"],
        "struck":["strike"],
        "string":["strung"],
        "strung":["string"],
        "strive":["strove","striven"],
        "strove":["strive","striven"],
        "striven":["strive","strove"],
        "swear":["swore","sworn"],
        "swore":["swear","sworn"],
        "sworn":["swore","swear"],
        "sweep":["swept"],
        "swept":["sweep"],
        "take":["took","taken"],
        "took":["take","taken"],
        "taken":["took","take"],
        "teach":["taught"],
        "taught":["teach"],
        "tear":["tore","torn"],
        "tore":["tear","torn"],
        "torn":["tear","tore"],
        "tell":["told"],
        "told":["tell"],
        "think":["thought"],
        "thought":["think"],
        "throw":["threw","thrown"],
        "threw":["throw","thrown"],
        "thrown":["throw","threw"],
        "tread":["treaded","trod","trodden"],
        "treaded":["tread","trod","trodden"],
        "trod":["treaded","tread","trodden"],
        "trodden":["treaded","trod","tread"],
        "understand":["understood"],
        "understood":["understand"],
        "undertake":["undertook","undertaken"],
        "undertook":["undertake","undertaken"],
        "undertaken":["undertake","undertook"],
        "undo":["undid","undone"],
        "undid":["undo","undone"],
        "undone":["undid","undo"],
        "wake":["woke"],
        "woke":["wake"],
        "wear":["wore","worn"],
        "wore":["wear","worn"],
        "worn":["wear","wore"],
        "weave":["wove","woven"],
        "wove":["weave","woven"],
        "woven":["wove","weave"],
        "weep":["wept"],
        "wept":["weep"],
        "win":["won"],
        "won":["win"],
        "wind":["wound"],
        "wound":["wind"],
        "withdraw":["withdrew","withdrawn"],
        "withdrew":["withdraw","withdrawn"],
        "withdrawn":["withdraw","withdrew"],
        "withstand":["withstood"],
        "withstood":["withstand"],
        "wring":["wrung"],
        "wrung":["wring"],
        "write":["wrote","written"],
        "wrote":["write","written"],
        "written":["write","wrote"]

    }
    graph = nx.read_adjlist(edgelist)
    # if statement to run this only when we use "null" in larger datasets
    if edgelist in ['words32768.adjlist', 'words65536.adjlist', 'words74286.adjlist', 'lemma_words32768.adjlist', 'lemma_words51228.adjlist']:
        # Missing word information ---------
        missing_word = "null"
        missing_word_IPA = ipa.convert(missing_word)
        missing_word_list = list(missing_word_IPA)
        graph.add_node(missing_word) 

        word_2_nlp = nlp(missing_word)
        for token in word_2_nlp:
            word_2_lemma = str(token.lemma_).strip()
        # --------------------------------------
        
        # Information to get similar data used previously. 
        words = []
        two_character_phonemes = ["oʊ", "ɔɪ", "aɪ", "aʊ"]
        vowels = ["ɑ", "æ", "ə", "ʌ", "ɔ", "a", "aɪ", "aʊ", "ɛ", "e", "ɪ", "i", "o", "ɔ", "ʊ", "u"]
        
       

        # Run through SUBTLEX and get the words and IPA translations
        df = pd.read_excel('SUBTLEX-US-Copy.xlsx')
        for j in range(len(df['Word'])):
            WORD = str(df['Word'][j]).strip()
            if WORD not in words.WORD:
                IPA = str(df['IPA'][j]).strip()
                IPA_LIST = str(df['IPA-List'][j]).strip().split()
                words.append(Words(WORD, IPA, IPA_LIST))

        # Look for the two phoneme rule, where they should as one character. This does not apply to null
        for i in range(0, len(words)):
            if words[i].WORD in G.nodes():
                word_1 = words[i].IPA_LIST
                temp_array_word = []
                temp_array_word.append(word_1[0])


                for k in range(1, len(word_1)):
                    temp_array_word.append(word_1[k])

                    if (word_1[k-1] + word_1[k] == "ər") and (k != len(word_1) - 1) and (word_1[k+1] not in vowels):
                            # print(f"{words[i].WORD}...... {word_1}.....{word_1[k+1]}")
                            temp_array_word.pop()
                            temp_array_word.pop()
                            temp_array_word.append(word_1[k-1] + word_1[k])


                    elif (word_1[k-1] + word_1[k] in two_character_phonemes):
                        temp_array_word.pop()
                        temp_array_word.pop()
                        temp_array_word.append(word_1[k-1] + word_1[k])
                
                words[i].IPA_LIST = temp_array_word
        
        
        # Create an edge between the nodes in the edgelists and null, and vice versa
        for k in range(0, len(words)):
            if words[k].WORD in G.nodes():
                word_1 = words[k].IPA_LIST
                word_1_nlp = nlp(words[k].WORD)
                for token in word_1_nlp:
                    word_1_lemma = str(token.lemma_).strip()

                # Check basic condition first
                if textdistance.levenshtein.distance(word_1, missing_word_list) == 1 and (word_2_lemma != word_1_lemma):
                    # Check if both are irregular
                    if (words[k].WORD not in irregular_words) or (missing_word not in irregular_words):
                        # return_string = return_string + words[j].WORD + " "
                        nx.add_edge(words[k].WORD, missing_word)
                        nx.add_edge(missing_word, words[k].WORD)
                        

                    # If both irregular then check irregular dictionary
                    elif missing_word not in irregular_words[words[k].WORD]:
                        # return_string = return_string + words[j].WORD + " "
                        nx.add_edge(words[k].WORD, missing_word)
                        nx.add_edge(missing_word, words[k].WORD)

        
    
    # Remove self loops
    for nodes in graph.nodes():
        if graph.has_edge(nodes,nodes):
            graph.remove_edge(nodes,nodes)
        
        if nodes in irregular_words:
            for irregular in irregular_words[nodes]:
                if graph.has_node(irregular) and graph.has_edge(nodes, irregular):
                    graph.remove_edge(nodes, irregular)  

    file = open(f"validate-{edgelist}", "wb")
    nx.write_adjlist(graph, file) 
    return graph



def getSystemFactsWordForms(edgeLists):
    table = []
    for i in edgeLists:
        print(i)
        graph = validateWordEdgeList(i)
        num_nodes = graph.number_of_nodes()
        num_edges = graph.number_of_edges()
        density = nx.density(graph)
        # giant_component = graph.subgraph(max(nx.connected_components(graph), key=len))
        
        all_subgraphs = sorted(nx.connected_components(graph), key=len, reverse=True)
        giant_component = graph.subgraph(all_subgraphs[0])
        number_of_islands = 0
        islands_nodes = 0
        hermits = 0

        nodes_in_giant_component = giant_component.number_of_nodes()
        for j in range(1,len(all_subgraphs)):
            if graph.subgraph(all_subgraphs[j]).number_of_nodes() != 1:
                islands_nodes += graph.subgraph(all_subgraphs[j]).number_of_nodes()
                number_of_islands += 1
            else:
                hermits += graph.subgraph(all_subgraphs[j]).number_of_nodes()
        shortest_path_length_giant_component = nx.average_shortest_path_length(giant_component) # <--- slows down calculation
        
        average_clustering = nx.average_clustering(giant_component)
        degree_assortaivity = nx.degree_assortativity_coefficient(graph)


        #Edit if validation works
        max_degree = 0
        total_degree = 0
        for k in graph.nodes():
            if graph.degree(k) >= max_degree:
                max_degree = graph.degree(k)

            # if graph.has_edge(k,k):
            #     total_degree += graph.degree(k) - 1
            total_degree += graph.degree(k)

        average_degree = (total_degree / num_nodes)
        hetrogeneity_parameter_k = math.sqrt(average_degree) / (average_degree * average_degree)
        table.append([num_nodes, num_edges, density, nodes_in_giant_component, islands_nodes, hermits, number_of_islands, shortest_path_length_giant_component, average_clustering, degree_assortaivity, max_degree, hetrogeneity_parameter_k])
    
    df = pd.DataFrame(table, columns = ['Nodes', 'Links', 'Density', 'Giant Nodes', 'Islands Nodes', 'Hermits Nodes', '# of Small Islands', 'Avg Short Path Giant', 'Average Clustering Coefficient Giant', 'Degree Assortativity', 'Max Degree', 'Heterogeneity Parameter K'], index=edgeLists)
    pd.set_option("display.max_columns", None)
    pd.set_option("display.max_rows", None)
    pd.set_option('display.width', 350)
    print(df)
    f = open("System-Facts-WF.txt", "w+")
    f.write(str(df))
    f.close()    


def getSystemFactsLemmas(edgeLists):
    table = []
    for i in edgeLists:
        print(i)
        graph = validateWordEdgeList(i)
        num_nodes = graph.number_of_nodes()
        num_edges = graph.number_of_edges()
        density = nx.density(graph)
        # giant_component = graph.subgraph(max(nx.connected_components(graph), key=len))
        
        all_subgraphs = sorted(nx.connected_components(graph), key=len, reverse=True)
        giant_component = graph.subgraph(all_subgraphs[0])
        number_of_islands = 0
        islands_nodes = 0
        hermits = 0

        nodes_in_giant_component = giant_component.number_of_nodes()
        for j in range(1,len(all_subgraphs)):
            if graph.subgraph(all_subgraphs[j]).number_of_nodes() != 1:
                islands_nodes += graph.subgraph(all_subgraphs[j]).number_of_nodes()
                number_of_islands += 1
            else:
                hermits += graph.subgraph(all_subgraphs[j]).number_of_nodes()
        shortest_path_length_giant_component = nx.average_shortest_path_length(giant_component) # <--- slows down calculation
        
        average_clustering = nx.average_clustering(giant_component)
        degree_assortaivity = nx.degree_assortativity_coefficient(graph)


        #Edit if validation works
        max_degree = 0
        total_degree = 0
        for k in graph.nodes():
            if graph.degree(k) >= max_degree:
                max_degree = graph.degree(k)

            # if graph.has_edge(k,k):
            #     total_degree += graph.degree(k) - 1
            total_degree += graph.degree(k)

        average_degree = (total_degree / num_nodes)
        hetrogeneity_parameter_k = math.sqrt(average_degree) / (average_degree * average_degree)
        
        table.append([num_nodes, num_edges, density, nodes_in_giant_component, islands_nodes, hermits, number_of_islands, shortest_path_length_giant_component, average_clustering, degree_assortaivity, max_degree, hetrogeneity_parameter_k])
    
    df = pd.DataFrame(table, columns = ['Nodes', 'Links', 'Density', 'Giant Nodes', 'Islands Nodes', 'Hermits Nodes', '# of Small Islands', 'Avg Shortset Path Giant', 'Average Clustering Coefficient Giant', 'Degree Assortativity', 'Max Degree', 'Heterogeneity Parameter K'], index=edgeLists)
    pd.set_option("display.max_columns", None)
    pd.set_option("display.max_rows", None)
    pd.set_option('display.width', 350)
    print(df)
    f = open("System-Facts-Lemma.txt", "w+")
    f.write(str(df))
    f.close()  


def getWordFacts(edgelist, list_of_words):
    graph = nx.read_adjlist(edgelist)
    y = 2
    excel_file = pd.read_excel(list_of_words)
    new_excel_file = excel_file
    new_excel_file.to_excel(f'Analyze-{words}', index = True, header = True)
    workbook = openpyxl.load_workbook(f'Analyze-{words}')
    worksheet_M = workbook.worksheets[0]
    worksheet_M.insert_cols(3)
    cell_title_M = worksheet_M.cell(row = 1, column = 3)
    cell_title_M.value = "Degree"

    worksheet_N = workbook.worksheets[0]
    worksheet_N.insert_cols(4)
    cell_title_N = worksheet_M.cell(row = 1, column = 4)
    cell_title_N.value = "Cluster Coefficient"

    df = pd.read_excel(f'Analyze-{words}')

    for i in range(len(df['Word'])):
        if str(df['Word'][i]).strip().lower() in graph.nodes():
            cell_to_write = worksheet_M.cell(row = y, column = 3)
            cell_to_write.value = graph.degree(str(df['Word'][i]).strip().lower()) 

            cell_to_write_ = worksheet_N.cell(row = y, column = 4)
            cell_to_write_.value = nx.clustering(graph, str(df['Word'][i]).strip().lower()) 
            y += 1
        
        elif str(df['Word'][i]).strip() in graph.nodes():
            cell_to_write = worksheet_M.cell(row = y, column = 3)
            cell_to_write.value = graph.degree(str(df['Word'][i]).strip())

            cell_to_write_ = worksheet_N.cell(row = y, column = 4)
            cell_to_write_.value = nx.clustering(graph, str(df['Word'][i]))  
            y += 1

        else:
            cell_to_write = worksheet_M.cell(row = y, column = 3)
            cell_to_write.value = "N/A"

            cell_to_write_ = worksheet_N.cell(row = y, column = 4)
            cell_to_write_.value = "N/A"
            y += 1

    workbook.save(f'Analyze-{words}')



# filename = input("Enter the path of your adjancey list: ")
# words = input("Enter the path of your xlsx word list: ")


# getWordFacts(filename, words)  


# getSystemFacts(['words1024.adjlist', 'words2048.adjlist', 'words4096.adjlist', 'words8192.adjlist', 'words16384.adjlist', 'words19839.adjlist', 'words32768.adjlist', 'words65536.adjlist', 'words74286.adjlist'])

# print("------------")
# getSystemFacts(['lemma_words1024.adjlist', 'lemma_words2048.adjlist', 'lemma_words4096.adjlist', 'lemma_words8192.adjlist', 'lemma_words16384.adjlist', 'lemma_words19839.adjlist', 'lemma_words32768.adjlist', 'lemma_words51228.adjlist'])

# graph = validate('words74286.adjlist')

getSystemFactsWordForms(['words1024.adjlist', 'words2048.adjlist', 'words4096.adjlist', 'words8192.adjlist', 'words16384.adjlist', 'words19839.adjlist', 'words32768.adjlist', 'words65536.adjlist', 'words74286.adjlist'])
print("\n")
getSystemFactsLemmas(['lemma_words1024.adjlist', 'lemma_words2048.adjlist', 'lemma_words4096.adjlist', 'lemma_words8192.adjlist', 'lemma_words16384.adjlist', 'lemma_words19839.adjlist', 'lemma_words32768.adjlist', 'lemma_words51228.adjlist'])



# df = pd.read_excel('SUBTLEX-US-Copy.xlsx')
#     # df = df.sample(frac = 0.0013)
#     # j = 0
#     # print(df['Word'])

#     # Original Code Below
# import collections 
# excel_file = []
# new_excel = []

# for i in range(len(df['Word'])):
#     WORD = str(df['Word'][i]).strip()
#     excel_file.append(WORD)

# # new_excel = list( dict.fromkeys(excel_file) ) 

# graph = nx.read_adjlist('words74286.adjlist')
# new_excel = list(graph.nodes())


# excel_file.sort()
# new_excel.sort()

# print(f"Length of excel: {len(excel_file)}")
# print(f"Length of excel: {len(new_excel)}")

# print([item for item, count in collections.Counter(excel_file).items() if count > 1])

# # The word null is missing. See if we run the adjancey lists BUT append this word here. 

# for i in range(0,len(excel_file)):
#     if excel_file[i] != new_excel[i]:
#         print(i)
#         print(f"{excel_file[i]} at excel and {new_excel[i]}")
#         print("--------")
#         print(i-2)
#         print(f"{excel_file[i - 2]} at excel and {new_excel[i - 2]}")
#         print("--------")
#         print("--------")
#         print(i-1)
#         print(f"{excel_file[i - 1]} at excel and {new_excel[i - 1]}")
#         print("--------")
#         print(i+1)
#         print(f"{excel_file[i + 1]} at excel and {new_excel[i + 1]}")
#         break


# print("\n \n \n")
# ma = 0
# for i in range(0,len(excel_file)):
#     if excel_file[i] in new_excel:
#         new_excel.remove(excel_file[i])
#     else:
#         ma = i

# print(excel_file[ma])
# print(len(new_excel))


# for i in (['nanas', 'nana']):
#     if i not in excel_file:
#         print("GGGGSG")

# print()


# print(excel_file[1021])
# print(new_excel[1021])


# file = open("validated-words74286.adjlist", "wb")
# nx.write_adjlist(graph, file) 
# graph2 = nx.read_adjlist('words250.adjlist')

# nx.draw(graph,
#  with_labels=True,
#  node_color='red',
#  node_size=8,
#  font_size=5,
#  verticalalignment='baseline',
#  edge_color='black')
# plt.show()


# nx.draw(graph2,
#  with_labels=True,
#  node_color='blue',
#  node_size=8,
#  font_size=5,
#  verticalalignment='baseline',
#  edge_color='black')
# plt.show()



